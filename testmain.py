# -*- coding: utf-8 -*-
import requests
import unittest
import json
from jinja2 import Template
from jinja2schema import infer, model

from oauthlib.oauth2 import LegacyApplicationClient
from oauthlib.oauth2 import BackendApplicationClient
from requests_oauthlib import OAuth2Session
from requests.auth import HTTPBasicAuth

import openpyxl
import sys

reload(sys)
sys.setdefaultencoding('utf8')

IP = 'http://localhost'
PORT = '8888'

ADDRESS = "%s:%s" % (IP, PORT)
URI = ""


def parse_excel(cases_file):
    user_cases = []
    wb = openpyxl.load_workbook(filename=cases_file)

    for name in wb.sheetnames:
        case = UserCase(name)
        ws = wb[name]
        current_step = None
        for row in ws.iter_rows(max_col=8):
            if row[0] == None and row[6] == None and row[7] == None:
                return
            if row[0].value == None:
                # if row[4].value <> None:
                #    current_step.request_data.append(row[4].value)
                if row[6].value <> None:
                    current_step.expect_value.append(row[6].value)
                if row[7].value <> None:
                    current_step.var_getvalue.append(row[7].value)
            else:
                step = Step(row[0].value, row[1].value, row[2].value, row[3].value, row[5].value)
                step.request_data.append(row[4].value)
                step.expect_value.append(row[6].value)
                step.var_getvalue.append(row[7].value)
            current_step = step
            case.steps.append(current_step)

        user_cases.append(case)
    return user_cases


def print_user_cases(user_cases):
    for case in user_cases:
        print "======================== case name =========================="
        print case.name
        for step in case.steps:
            print "=====step case======="
            print "step_name=", step.step_name
            print "request_type=", step.request_type
            print "request_url=", step.request_url
            print "request_data_type=", step.request_data_type
            print "expect_code=", step.expect_code
            print "request_data=", step.request_data
            print "expect_value=", step.expect_value
            print "var_getvalue=", step.var_getvalue
            print "===================="
        print "==============================================="


class UserCase:
    def __init__(self, name):
        headers_data = TestOauthRequest()
        authorization = headers_data[u'token_type'] + " " + headers_data[u'access_token']
        headers = {
            "content-type": "application/json",
            "tenant": "tenant_test",
            "authorization": authorization,
        }
        self.headers = headers
        self.name = name
        self.steps = []
        self.env = {}


class Step:
    def __init__(self, case_name, request_type, request_url, request_data_type, expect_code):
        self.step_name = case_name
        self.request_type = request_type
        self.request_url = request_url
        self.request_data_type = request_data_type
        self.request_data = []
        self.expect_code = expect_code
        self.expect_value = []
        self.var_getvalue = []


class TestEngine:
    def __init__(self):
        '''
        data = {}
        headers_data = TestOauthRequest()
        authorization = headers_data[u'token_type'] + " " + headers_data[u'access_token']
        headers = {
            'content-type': "application/json",
            'tenant': "tenant_test",
            'authorization': authorization,
        }
        self.headers = headers
        #self.request_list["put"] = TestPutRequest
        #self.request_list["get"] = TestGetRequest
        #self.request_list["delete"] = TestDeleteRequest
        '''
        self.request_list = {}
        self.request_list["post"] = TestPostRequest
        self.request_list["put"] = TestPutRequest
        self.request_list["get"] = TestGetRequest
        self.request_list["delete"] = TestDeleteRequest
        self.request_list["patch"] = TestPatchRequest
        self.response_data = {}

    def execute(self, storys):
        for cases in storys:
            self.docase(cases)

    def docase(self, cases):
        for case in cases.steps:
            self.step_execute(cases, case, case.step_name, case.request_type, case.request_url, case.request_data_type,
                              case.request_data,
                              case.expect_code, case.expect_value, case.var_getvalue)

    def step_execute(self, cases, case, step_name, request_type, request_url, request_data_type, request_data,
                     expect_result, expect_value,
                     var_getvalue):
        if len(infer(request_data[0]).items()) <> 0:
            template = Template(request_data[0])
            key = infer(request_data[0]).items()[0][0]
            t = dict()
            t.__setitem__(key, cases.env[key])
            request_data[0] = template.render(t)
            # print request_data[0]
        if len(infer(request_url).items()) <> 0:
            template = Template(request_url)
            key = infer(request_url).items()[0][0]
            t = dict()
            if key in cases.env:
                t.__setitem__(key, cases.env[key])
                request_url = template.render(t)
                # print request_url

        if request_type == None:
            return
        # status, content = self.request_list[request_type.encode('ascii')](request_url, "path", "test")
        status, content = self.request_list[request_type](request_url, cases.headers, request_data_type, request_data)
        # print "status=", status
        # print "content:", content
        content_data = json.loads(content)
        print content_data
        if status != 200:
            error = content_data[u'message'].encode('utf-8')
            print "\033[31;1m[FAILED] %s" % step_name + "请求失败，错误信息为：" + str(status) + "," + error
            return
        # print(content_data[u'result'])
        # print(content_data[u'result'] == expect_result)
        if 200 == status and content_data[u'result'] == expect_result:
            print "\033[32;1m[SUCCESS] %s" % step_name
            # 获取出参
            print(case.var_getvalue)
            for val in case.var_getvalue:
                if val == "{}" or val == "" or val == None:
                    k+=1
                    continue
                t = eval(val)
                for tt in t:
                    result = self.get_content(content, t[tt])
                  #  cases.env[t.items()[0][0]] = result
                    cases.env[tt] = result
                # print(cases.env)
            # 检查入参

            for val in case.expect_value:
                if val == "{}" or val == "" or val == None:
                    continue
                t = eval(val)
                if self.check_content(content, t.items()[0][1], t.items()[0][0]):
                    print "\033[32;1m[SUCCESS] %s" % t.items()[0][0]
                else:
                    print "\033[31;1m[FAILED] %s" % t.items()[0][0]
            print cases.env
        else:
            print "\033[31;1m[FAILED] %s" % step_name + content_data[u'message']
        print request_url

    # key = [key1, key2, key3]
    def get_content(self, content, json_keys):
        result_json = json.loads(content)
        i = 0
        temp_json = result_json
        for jkey in json_keys:
            if jkey in temp_json:
                temp_json = temp_json[jkey]
                i = i + 1
            elif type(jkey) == type(1):
                temp_json = temp_json[jkey]
                i = i + 1
            else:
                break
        if len(json_keys) == i:
            return temp_json
        else:
            return ""

    def check_content(self, content, json_keys, json_value):
        result_json = json.loads(content)
        i = 0
        temp_json = result_json
        for jkey in json_keys:
            if jkey in temp_json:
                temp_json = temp_json[jkey]
                i = i + 1
            elif type(jkey) == type(1):
                temp_json = temp_json[jkey]
                i = i + 1
            else:
                break
        if len(json_keys) == i and temp_json == json_value:
            return True
        else:
            return False


'''
def TestRequest(type, url, reuest_data, headers):
    body = None
    params = None
    if "params" in reuest_data:
        params = reuest_data["params"]
    if "body" in reuest_data:
        body = reuest_data["body"]
    print(ADDRESS + url)
    print(body)
    print(params)
    if params == None and body == None:
        r = requests.request(type, ADDRESS + url, headers=headers)
    elif params != None and body != None:
        r = requests.request(type, ADDRESS + url, params=params, data=body, headers=headers)
    elif params != None and body == None:
        r = requests.request(type, ADDRESS + url, params=params, headers=headers)
    elif params == None and body != None:
        r = requests.request(type, ADDRESS + url, headers=headers,data=body)
    return r;
'''


def TestGetRequest(url, headers, data_type, data):
    url = ADDRESS + url
    if data[0] == None:
        r = requests.get(url, headers=headers)
    else:
        request_data = data[0].encode("utf-8")
        if data_type == "path":
            r = requests.get(url, headers=headers, params=request_data)
        elif data_type == "body":
            r = requests.get(url, headers=headers, data=request_data)

    return r.status_code, r.content


def TestPostRequest(url, headers, data_type, data):
    # return 200, '{"test": {"test2": {"test": "value"} } }'
    url = ADDRESS + url
    request_data = data[0].encode("utf-8")
    if data == None:
        r = requests.post(url,headers=headers)
    else:
        if data_type == "path":
            r = requests.post(url, headers=headers, params=request_data)
        elif data_type == "body":
            # data[0]='{"deletable":"1","description":"测试2角1色","name":"测试角色1181","type":"自定义"}'
            r = requests.post(url, headers=headers, data=request_data)

    return r.status_code, r.content


def TestPutRequest(url, headers, data_type, data):
    url = ADDRESS + url
    request_data = data[0].encode("utf-8")
    if data[0] == None:
        r = requests.put(url, headers=headers)
    else:
        if data_type == "path":
            r = requests.put(url, headers=headers, params=request_data)
        elif data_type == "body":
            r = requests.put(url, headers=headers, data=request_data)

    return r.status_code, r.content


def TestPatchRequest(url, headers, data_type, data):
    url = ADDRESS + url
    if data[0] == None:
        r = requests.patch(url, headers=headers)
    else:
        request_data = data[0].encode("utf-8")
        if data_type == "path":
            r = requests.patch(url, headers=headers, params=request_data)
        elif data_type == "body":
            r = requests.patch(url, headers=headers, data=request_data)
    return r.status_code, r.content


def TestDeleteRequest(url, headers, data_type, data):
    url = ADDRESS + url
    if data[0] == None:
        r = requests.delete(url, headers=headers)
    else:
        request_data = data[0].encode("utf-8")
        if data_type == "path":
            r = requests.delete(url, headers=headers, params=request_data)
        elif data_type == "body":
            r = requests.delete(url, headers=headers, data=request_data)

    return r.status_code, r.content


def ExpectResult():
    pass


# 获取token信息
def TestOauthRequest():
    r = requests.post(ADDRESS + "/oauth/token?grant_type=password&username=user&password=password",
                      auth=HTTPBasicAuth('chiron-client', 'chiron'))
    token_data = json.loads(r.content)
    print r.content
    r = requests.post(ADDRESS + "/oauth/token?grant_type=refresh_token&refresh_token=%s" % token_data[u'refresh_token'],
                      auth=HTTPBasicAuth('', ''))
    return token_data


if __name__ == '__main__':
    # print TestOauthRequest()
    # print(TestOauthRequest())

    # te = TestEngine()
    # for index,value in enumerate(testcase.cases):
    #    te.addcase(value)
    #    te.execute()

    # print te.check_content('{"test1": {"test2": 1}}', ["test1", "test2"], 1)
    print ADDRESS
    user_cases = parse_excel("testcase.xlsx")
    print_user_cases(user_cases)
    te = TestEngine()
    te.execute(user_cases)
